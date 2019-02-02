# HW4 Section 1 (Qplayer) self play and Qplayer vs Human player. Final version.

import numpy as np
import tkinter as tk
import copy
import pickle as pickle  

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'example.xlsx'
ws1 = wb.active
ws1.title = "Training Results"

i=1

from Q_Learning_Tic_Tac_Toe import Game, HumanPlayer, QPlayer, Board
board = Board()

Q = pickle.load(open("Q_epsilon_09_Nepisodes_20000.p", "rb"))

root = tk.Tk()
player1 = QPlayer(mark="X", epsilon=0.01)
player2 = QPlayer(mark="O", epsilon=0.01)

game = Game(root, player1, player2, Q=Q)

N_episodes = 20000

for episodes in range(N_episodes):
    game.play()
    game.reset()

    if game.board.winner() is None:
        ws1.cell(row=i+1, column=1, value=0)
    else:
        if game.current_player.mark == "X":
            ws1.cell(row=i+1, column=1, value=1)
        elif game.current_player.mark == "O":
            ws1.cell(row=i+1, column=1, value=-1)
    i += 1

#game.play()
#root.mainloop()

wb.save(filename = dest_filename)