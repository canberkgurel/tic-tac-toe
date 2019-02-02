# Q-learning-Tic-Tac-Toe
# Canberk Final Version HW4 Section 1 Training 

import numpy as np
import tkinter as tk
import copy
import pickle
from Q_Learning_Tic_Tac_Toe import Game, QPlayer, Board     # Classes used for Tic Tac Toe

board = Board()
'''
import xlwt
from datetime import datetime

wb = xlwt.Workbook()
ws = wb.add_sheet('Training Results')
'''
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'example.xlsx'
ws1 = wb.active
ws1.title = "Training Results"

i=1

root = tk.Tk()
epsilon = 0.9
player1 = QPlayer(mark="X",epsilon = epsilon)
player2 = QPlayer(mark="O",epsilon = epsilon)
game = Game(root, player1, player2)

N_episodes = 200000

for episodes in range(N_episodes):
    game.play()
    game.reset()

    if game.board.winner() is None:
        #ws.write(i+1, 0, 0)
        #ws1.append([0])
        ws1.cell(row=i+1, column=1, value=0)
    else:
        if game.current_player.mark == "X":
            #ws.write(i+1, 0, 1)
            #ws1.append([1])
            ws1.cell(row=i+1, column=1, value=1)
        elif game.current_player.mark == "O":
            #ws.write(i+1, 0, -1) 
            #ws1.append([0])
            ws1.cell(row=i+1, column=1, value=-1)
    i += 1

Q = game.Q

filename = "Q_epsilon_09_Nepisodes_{}.p".format(N_episodes)
pickle.dump(Q, open(filename, "wb"))

wb.save(filename = dest_filename)
